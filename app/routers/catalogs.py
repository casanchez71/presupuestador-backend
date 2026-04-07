"""Catalog price lookup and application to budgets."""

from __future__ import annotations

import csv
import io
import warnings
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile

from app.auth import get_current_user
from app.db import get_data_db
from app.schemas import CatalogTipo

router = APIRouter()

# ── Tab-name → tipo mapping ───────────────────────────────────────────────────

TAB_TIPO_MAP: dict[str, str] = {
    "materiales": "material",
    "material": "material",
    "mat": "material",
    "mano de obra": "mano_obra",
    "mano_obra": "mano_obra",
    "mo": "mano_obra",
    "equipos": "equipo",
    "equipo": "equipo",
    "eq": "equipo",
    "subcontratos": "subcontrato",
    "subcontrato": "subcontrato",
    "sub": "subcontrato",
}

# Flexible column aliases for price column
_PRICE_ALIASES = {"precio_unitario", "precio_sin_iva", "precio", "costo", "precio_unit", "p_unitario"}


# ── Upload CSV catalog ────────────────────────────────────────────────────


@router.post("/upload-csv")
async def upload_csv_catalog(
    file: UploadFile = File(...),
    tipo: CatalogTipo = Query(..., description="Tipo: material, mano_obra, equipo, subcontrato"),
    name: str = Query(None, description="Nombre del catalogo (default: nombre del archivo)"),
    user: dict = Depends(get_current_user),
):
    """Upload a CSV price list and create a catalog with entries.

    CSV must have columns: codigo, descripcion, unidad, precio_unitario
    """
    db = get_data_db()
    org_id = user["org_id"]

    # Read and parse CSV
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")  # handle BOM
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))

    # Validate required columns
    required_cols = {"codigo", "descripcion", "unidad", "precio_unitario"}
    if not reader.fieldnames or not required_cols.issubset({c.strip().lower() for c in reader.fieldnames}):
        raise HTTPException(
            400,
            f"CSV debe tener columnas: {', '.join(sorted(required_cols))}. "
            f"Encontradas: {reader.fieldnames}",
        )

    # Normalize fieldnames
    field_map = {c.strip().lower(): c for c in reader.fieldnames}

    rows = []
    for row in reader:
        codigo = (row.get(field_map.get("codigo", "codigo")) or "").strip()
        descripcion = (row.get(field_map.get("descripcion", "descripcion")) or "").strip()
        unidad = (row.get(field_map.get("unidad", "unidad")) or "").strip()
        precio_raw = (row.get(field_map.get("precio_unitario", "precio_unitario")) or "").strip()

        if not codigo or not precio_raw:
            continue

        try:
            # Handle Argentine format (dot as thousands, comma as decimal)
            precio_clean = precio_raw.replace(".", "").replace(",", ".") if "," in precio_raw else precio_raw
            precio = float(precio_clean)
        except ValueError:
            continue

        rows.append({
            "codigo": codigo,
            "descripcion": descripcion,
            "unidad": unidad,
            "precio_sin_iva": precio,
            "tipo": tipo,
        })

    if not rows:
        raise HTTPException(400, "CSV sin filas validas")

    # Create catalog
    catalog_name = name or (file.filename or "catalogo").rsplit(".", 1)[0]
    catalog_result = db.table("price_catalogs").insert({
        "org_id": org_id,
        "name": catalog_name,
        "source_file": file.filename,
    }).execute()
    if not catalog_result.data:
        raise HTTPException(500, "Error al crear catalogo")
    catalog_id = catalog_result.data[0]["id"]

    # Insert entries
    entries = [
        {
            "catalog_id": catalog_id,
            "org_id": org_id,
            **row,
        }
        for row in rows
    ]
    db.table("catalog_entries").insert(entries).execute()

    return {
        "catalog_id": catalog_id,
        "name": catalog_name,
        "entries_count": len(entries),
        "tipo": tipo,
    }


# ── Upload Excel catalog (multi-tab) ─────────────────────────────────────────


def _parse_excel_rows(ws) -> list[dict]:  # type: ignore[no-untyped-def]
    """Extract rows from an openpyxl worksheet using flexible column aliases."""
    data_rows = list(ws.iter_rows(values_only=True))
    if not data_rows:
        return []

    # Find header row (first row with at least one non-None value)
    header_idx = 0
    headers: list[str] = []
    for idx, row in enumerate(data_rows):
        if any(v is not None for v in row):
            headers = [str(v).strip().lower() if v is not None else "" for v in row]
            header_idx = idx
            break

    if not headers:
        return []

    # Map normalized header → column index (first occurrence wins)
    field_map: dict[str, int] = {}
    for col_idx, h in enumerate(headers):
        if h and h not in field_map:
            field_map[h] = col_idx

    def find_col(candidates: set[str]) -> int | None:
        for c in candidates:
            if c in field_map:
                return field_map[c]
        return None

    codigo_col = find_col({"codigo", "cod", "code"})
    descripcion_col = find_col({"descripcion", "descripción", "description", "nombre", "name"})
    unidad_col = find_col({"unidad", "unit", "ud"})
    precio_col = find_col(_PRICE_ALIASES)

    if descripcion_col is None or precio_col is None:
        return []

    def _cell(row: tuple, idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return str(v).strip() if v is not None else ""

    rows: list[dict] = []
    for row in data_rows[header_idx + 1:]:
        if not any(v is not None for v in row):
            continue  # skip blank rows

        codigo = _cell(row, codigo_col) if codigo_col is not None else ""
        descripcion = _cell(row, descripcion_col)
        unidad = _cell(row, unidad_col) if unidad_col is not None else ""
        precio_raw = _cell(row, precio_col)

        if not descripcion or not precio_raw:
            continue

        try:
            precio_clean = (
                precio_raw.replace(".", "").replace(",", ".")
                if "," in precio_raw
                else precio_raw
            )
            precio = float(precio_clean)
        except (ValueError, AttributeError):
            continue

        rows.append({
            "codigo": codigo,
            "descripcion": descripcion,
            "unidad": unidad,
            "precio_sin_iva": precio,
        })

    return rows


@router.post("/upload-excel")
async def upload_excel_catalog(
    file: UploadFile = File(...),
    name: str = Query(None, description="Prefijo de nombre para los catalogos (default: nombre del archivo)"),
    user: dict = Depends(get_current_user),
):
    """Upload an Excel file (.xlsx/.xls) with up to 4 tabs and create one catalog per tab.

    Tab name matching (case-insensitive):
    - Materiales / Material / Mat → tipo material
    - Mano de obra / mano_obra / MO → tipo mano_obra
    - Equipos / Equipo / Eq → tipo equipo
    - Subcontratos / Subcontrato / Sub → tipo subcontrato

    Each tab must have columns: codigo, descripcion, unidad + a price column
    (flexible aliases accepted: precio_unitario, precio_sin_iva, precio, costo, etc.)
    """
    if file.filename and not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "El archivo debe ser .xlsx o .xls")

    try:
        import openpyxl  # type: ignore[import]
    except ImportError:
        raise HTTPException(500, "openpyxl no esta instalado en el servidor")

    content = await file.read()
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"No se pudo leer el archivo Excel: {exc}")

    db = get_data_db()
    org_id = user["org_id"]
    base_name = name or (file.filename or "catalogo").rsplit(".", 1)[0]

    catalogs_created = 0
    entries_summary: dict[str, int] = {}
    warnings_list: list[str] = []

    for sheet_name in wb.sheetnames:
        tipo = TAB_TIPO_MAP.get(sheet_name.strip().lower())
        if tipo is None:
            warnings_list.append(f"Solapa '{sheet_name}' ignorada — nombre no reconocido")
            continue

        ws = wb[sheet_name]
        try:
            rows = _parse_excel_rows(ws)
        except Exception as exc:
            warnings_list.append(f"Solapa '{sheet_name}' con error al leer: {exc}")
            continue

        if not rows:
            warnings_list.append(f"Solapa '{sheet_name}' sin filas validas — saltada")
            continue

        # Create catalog
        catalog_name = f"{base_name} - {sheet_name}"
        catalog_result = db.table("price_catalogs").insert({
            "org_id": org_id,
            "name": catalog_name,
            "source_file": file.filename,
        }).execute()

        if not catalog_result.data:
            warnings_list.append(f"No se pudo crear el catalogo para la solapa '{sheet_name}'")
            continue

        catalog_id = catalog_result.data[0]["id"]

        # Insert entries with tipo
        entries = [
            {
                "catalog_id": catalog_id,
                "org_id": org_id,
                "tipo": tipo,
                **row,
            }
            for row in rows
        ]
        db.table("catalog_entries").insert(entries).execute()

        catalogs_created += 1
        entries_summary[tipo] = entries_summary.get(tipo, 0) + len(entries)

    wb.close()

    if catalogs_created == 0:
        raise HTTPException(
            400,
            "No se creo ningun catalogo. Verificá que las solapas se llamen: "
            "Materiales, Mano de obra, Equipos, Subcontratos (o variantes como Mat, MO, Eq, Sub).",
        )

    return {
        "catalogs_created": catalogs_created,
        "entries": entries_summary,
        "warnings": warnings_list,
        "source_file": file.filename,
    }


# ── List catalogs ───────────────────────────────────────────────────────────


@router.get("")
async def list_catalogs(user: dict = Depends(get_current_user)):
    """List all price catalogs for the org."""
    db = get_data_db()
    result = (
        db.table("price_catalogs")
        .select("*")
        .eq("org_id", user["org_id"])
        .order("created_at", desc=True)
        .execute()
    )
    return result.data or []


# ── List catalog entries ────────────────────────────────────────────────────


@router.get("/{catalog_id}/entries")
async def list_catalog_entries(
    catalog_id: UUID,
    tipo: str | None = Query(None, description="Filtrar por tipo: material, mano_obra, equipo, subcontrato"),
    user: dict = Depends(get_current_user),
):
    """List entries in a catalog, with optional tipo filter."""
    db = get_data_db()
    org_id = user["org_id"]
    cid = str(catalog_id)

    # Verify catalog belongs to org
    catalog = (
        db.table("price_catalogs")
        .select("id")
        .eq("id", cid)
        .eq("org_id", org_id)
        .single()
        .execute()
    )
    if not catalog.data:
        raise HTTPException(404, "Catalogo no encontrado")

    q = (
        db.table("catalog_entries")
        .select("*")
        .eq("catalog_id", cid)
        .eq("org_id", org_id)
    )
    if tipo:
        q = q.eq("tipo", tipo)

    result = q.order("codigo").execute()
    return result.data or []


# ── Search catalog entries ──────────────────────────────────────────────────


@router.get("/{catalog_id}/search")
async def search_catalog_entries(
    catalog_id: UUID,
    q: str = Query(..., min_length=1, description="Texto de busqueda en descripcion"),
    user: dict = Depends(get_current_user),
):
    """Search catalog entries by description (case-insensitive)."""
    db = get_data_db()
    org_id = user["org_id"]
    cid = str(catalog_id)

    # Verify catalog belongs to org
    catalog = (
        db.table("price_catalogs")
        .select("id")
        .eq("id", cid)
        .eq("org_id", org_id)
        .single()
        .execute()
    )
    if not catalog.data:
        raise HTTPException(404, "Catalogo no encontrado")

    result = (
        db.table("catalog_entries")
        .select("*")
        .eq("catalog_id", cid)
        .eq("org_id", org_id)
        .ilike("descripcion", f"%{q}%")
        .execute()
    )
    return result.data or []


# ── Apply catalog prices to budget ──────────────────────────────────────────


@router.post("/apply/{budget_id}/{catalog_id}")
async def apply_catalog_to_budget(
    budget_id: UUID,
    catalog_id: UUID,
    user: dict = Depends(get_current_user),
):
    """Apply catalog prices to budget item_resources by matching codigo.

    For each item_resource in the budget, look up its price from the catalog
    (match by codigo) and update precio_unitario. Recalculate subtotals.
    """
    db = get_data_db()
    org_id = user["org_id"]
    bid = str(budget_id)
    cid = str(catalog_id)

    # Verify budget belongs to org
    budget = (
        db.table("budgets")
        .select("id")
        .eq("id", bid)
        .eq("org_id", org_id)
        .single()
        .execute()
    )
    if not budget.data:
        raise HTTPException(404, "Presupuesto no encontrado")

    # Verify catalog belongs to org
    catalog = (
        db.table("price_catalogs")
        .select("id")
        .eq("id", cid)
        .eq("org_id", org_id)
        .single()
        .execute()
    )
    if not catalog.data:
        raise HTTPException(404, "Catalogo no encontrado")

    # Load all catalog entries indexed by codigo
    entries_result = (
        db.table("catalog_entries")
        .select("codigo, precio_sin_iva, tipo")
        .eq("catalog_id", cid)
        .eq("org_id", org_id)
        .execute()
    )
    price_map: dict[str, float] = {}
    for entry in (entries_result.data or []):
        codigo = (entry.get("codigo") or "").strip()
        precio = entry.get("precio_sin_iva")
        if codigo and precio is not None:
            price_map[codigo] = float(precio)

    if not price_map:
        raise HTTPException(404, "Catalogo sin entradas con precios")

    # Get all budget items
    items = (
        db.table("budget_items")
        .select("id")
        .eq("budget_id", bid)
        .eq("org_id", org_id)
        .execute()
    )
    if not items.data:
        raise HTTPException(404, "Presupuesto sin items")

    item_ids = [item["id"] for item in items.data]

    # Get all item_resources for these items
    all_resources: list[dict] = []
    for item_id in item_ids:
        res = (
            db.table("item_resources")
            .select("*")
            .eq("item_id", item_id)
            .eq("org_id", org_id)
            .execute()
        )
        all_resources.extend(res.data or [])

    if not all_resources:
        raise HTTPException(404, "Presupuesto sin recursos en items")

    matched = 0
    unmatched = 0
    total_updated = 0.0

    for resource in all_resources:
        codigo = (resource.get("codigo") or "").strip()
        if codigo not in price_map:
            unmatched += 1
            continue

        new_precio = price_map[codigo]
        cantidad_eff = resource.get("cantidad_efectiva") or resource.get("cantidad") or 0
        new_subtotal = round(new_precio * float(cantidad_eff), 2)

        db.table("item_resources").update({
            "precio_unitario": new_precio,
            "subtotal": new_subtotal,
        }).eq("id", resource["id"]).execute()

        matched += 1
        total_updated += new_subtotal

    # Recalculate budget_items totals from their resources
    for item_id in item_ids:
        resources = (
            db.table("item_resources")
            .select("tipo, subtotal")
            .eq("item_id", item_id)
            .eq("org_id", org_id)
            .execute()
        )
        mat_total = 0.0
        mo_total = 0.0
        for r in (resources.data or []):
            subtotal = r.get("subtotal") or 0
            if r.get("tipo") == "material":
                mat_total += subtotal
            elif r.get("tipo") == "mano_obra":
                mo_total += subtotal

        directo_total = mat_total + mo_total
        # Preserve existing indirecto and beneficio
        item_data = (
            db.table("budget_items")
            .select("indirecto_total, beneficio_total, cantidad")
            .eq("id", item_id)
            .single()
            .execute()
        )
        if not item_data.data:
            continue

        cantidad = item_data.data.get("cantidad") or 0
        indirecto = item_data.data.get("indirecto_total") or 0
        beneficio = item_data.data.get("beneficio_total") or 0
        neto_total = directo_total + indirecto + beneficio

        mat_unitario = round(mat_total / cantidad, 2) if cantidad else 0
        mo_unitario = round(mo_total / cantidad, 2) if cantidad else 0

        db.table("budget_items").update({
            "mat_unitario": mat_unitario,
            "mo_unitario": mo_unitario,
            "mat_total": round(mat_total, 2),
            "mo_total": round(mo_total, 2),
            "directo_total": round(directo_total, 2),
            "neto_total": round(neto_total, 2),
        }).eq("id", item_id).execute()

    return {
        "items_matched": matched,
        "items_unmatched": unmatched,
        "total_updated": round(total_updated, 2),
    }
